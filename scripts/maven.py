import subprocess
import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from scripts import constants

EXCEL_ERROR_FILE = constants.ERROR_XLSX

def registrar_error_en_excel(ruta_proyecto, mensaje_error, detalle):
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    headers = ["Fecha", "Ruta del Proyecto", "Mensaje de Error", "Detalles"]

    ruta_proyecto = os.path.normpath(os.path.abspath(ruta_proyecto))  # Normaliza para evitar duplicados

    if not os.path.exists(EXCEL_ERROR_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores"
        ws.append(headers)
    else:
        wb = load_workbook(EXCEL_ERROR_FILE)
        ws = wb.active

        # Verificar si ya existe este proyecto registrado
        for row in ws.iter_rows(min_row=2, values_only=True):
            if os.path.normpath(row[1]) == ruta_proyecto:
                print(f"⚠️ Ya existe un error registrado para: {ruta_proyecto}")
                return  # Ya está registrado

    ws.append([fecha, ruta_proyecto, mensaje_error, detalle[:30000]])

    # Ajustar anchos solo si es la primera fila
    if ws.max_row == 2:
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(20, len(header) + 2)

    wb.save(EXCEL_ERROR_FILE)
    print(f"📝 Error registrado en {EXCEL_ERROR_FILE}")

def build_maven_project(project_dir):
    print(f"🔧 Compilando: {project_dir}")

    if not os.path.isdir(project_dir):
        print(f"❌ La ruta no existe o no es un directorio: {project_dir}")
        registrar_error_en_excel(project_dir, "Ruta inválida", "No es un directorio o no existe")
        return

    try:
        result = subprocess.run(
            ["mvn", "clean", "install", "-DskipTests", "-Dorg.slf4j.simpleLogger.defaultLogLevel=error"],
            cwd=project_dir,
            shell=True,
            text=True,
            capture_output=True
        )

        if result.returncode != 0:
            print("❌ Maven falló.")
            registrar_error_en_excel(
                project_dir,
                "Error de compilación Maven",
                f"STDOUT:\n{result.stdout or 'Sin salida'}\n\nSTDERR:\n{result.stderr or 'Sin error'}"
            )
        else:
            print("✅ Maven ejecutado correctamente.")
            print(result.stdout)

    except FileNotFoundError:
        print("❌ El comando 'mvn' no está disponible.")
        registrar_error_en_excel(
            project_dir,
            "Maven no encontrado",
            "Verifica que 'mvn' esté instalado y en el PATH del sistema."
        )

    except Exception as ex:
        print(f"❌ Error inesperado: {ex}")
        registrar_error_en_excel(project_dir, "Error inesperado", str(ex))

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python build_project.py <ruta_a_proyecto>")
    else:
        build_maven_project(sys.argv[1])
