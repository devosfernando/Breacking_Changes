import subprocess
import os
import re
import json
import constants
from openpyxl import Workbook, load_workbook

result = []
def parse_bitbucket_url(job,version,artifact):
    nueva_version = re.sub(r'^(\d+\.\d+)\.0$', r'\1', version)
    branch = "tag/"+nueva_version

    ruta = artifact
    dsg = ""
    match = re.search(r'/([^/]+)/[^/]+\.jar$', ruta)
    if match:
        dsg = match.group(1)

    repo_name = job

    project = dsg
    repo = repo_name.lower()

    
    git_url = f"https://bitbucket.globaldevtools.bbva.com/bitbucket/scm/{project}/{repo}.git"
    return git_url, branch, repo_name

def clone_repo(repo_urls, destination_dir=constants.SOURCE):
    failed_repos = []
    error_file = constants.CLONE_ERROR_XLSX
    for repo in repo_urls:
        try:
            job = repo["job"]
            version = repo["version"]
            artifact = repo["artifact"]

            git_url, branch, repo_name = parse_bitbucket_url(job, version, artifact)

            print(f"🧩 Clonando repo: {git_url} en la rama: {branch}")
            repo_path = os.path.join(destination_dir, repo_name)
            if not os.path.exists(repo_path):
                subprocess.check_call(["git", "clone", git_url, repo_path])
                subprocess.run(["git", "checkout", f"tags/{version}"], cwd=repo_path, check=True)
                print("✅ Clonado exitoso.")
            else:
                print(f"⚠️ Ya existe {repo_path}, se omite clonación.")

        except Exception as e:
            error_msg = str(e)
            print(f"❌ Error al clonar {repo.get('job')}: {error_msg}")
            failed_repos.append([repo.get("job"), repo.get("version"), repo.get("artifact"), error_msg])


    # Guardar errores en Excel
    if failed_repos:
        if os.path.exists(error_file):
            # Si ya existe, lo cargamos y agregamos filas
            wb = load_workbook(error_file)
            ws = wb.active
        else:
            # Crear uno nuevo con encabezados
            wb = Workbook()
            ws = wb.active
            ws.append(["Job", "Version", "Artifact", "Error"])

        for row in failed_repos:
            ws.append(row)

        wb.save(error_file)
        print(f"📊 Errores registrados en {error_file}")

# Ejemplo de uso:
# if __name__ == "__main__":
def clone():    
    with open(constants.PRODUCTIVE_JSON, "r", encoding="utf-8") as f:
        datos = json.load(f)
        clone_repo(datos)
