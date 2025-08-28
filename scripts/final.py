import os
import shutil
import xml.etree.ElementTree as ET
import re
import sys
from openpyxl import load_workbook, Workbook
from scripts import maven
import stat
from scripts import constants

file_path = constants.VERSIONS_XLSX

def copiar_contenido(origen):
    output_base = constants.MODIFIED
    nombre = os.path.basename(origen.rstrip("/\\"))
    destino = os.path.join(output_base, nombre)

    if os.path.exists(destino):
        shutil.rmtree(destino)

    shutil.copytree(origen, destino)
    return destino

def update_pom_in_jar(dir_path, new_parent_version, pom_filename):
    pom_path = None
    for root_dir, _, files in os.walk(dir_path):
        for file in files:
            if file == 'pom.xml':
                pom_path = os.path.join(root_dir, file)
                break

    if not pom_path:
        print(f"❌ No se encontró pom.xml en: {dir_path}")
        return

    tree = ET.parse(pom_path)
    root = tree.getroot()
    ns_uri = root.tag.split('}')[0].strip('{')
    ns = {'m': ns_uri}

    # Excel
    archivo_existente = os.path.exists(file_path)
    if archivo_existente:
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cambios de versión"
        ws.append(["artifactId", "old_version", "new_version", "archivo"])

    parent_node = root.find('m:parent', ns)
    artifact = root.find('m:artifactId', ns)
    if parent_node is not None:
        version_node = parent_node.find('m:version', ns)
        if version_node is not None:
            old = version_node.text
            version_node.text = new_parent_version
            print(f"🔧 Cambiando versión del parent: {old} → {new_parent_version} | job → {artifact.text}")
            ws.append([artifact.text, old, new_parent_version, pom_filename])
        else:
            print("⚠️ No se encontró <version> dentro de <parent>")
    else:
        print("⚠️ No se encontró <parent> en el pom")

    ET.register_namespace('', ns_uri)
    tree.write(pom_path, encoding="utf-8", xml_declaration=True)
    wb.save(file_path)

    return pom_path

def listar_jars_recursivo(target_pom_version,directorio, modificar=False):
    if not os.path.isdir(directorio):
        print(f"❌ Directorio inválido: {directorio}")
        return []

    encontrados = []
    pom_paths = []

    # Primero recolectar todos los pom.xml para saber cuántos hay
    for root, _, files in os.walk(directorio):
        for file in files:
            if file.endswith("pom.xml"):
                ruta_pom = os.path.join(root, file).replace("\\", "/")
                pom_paths.append(ruta_pom)

    total = len(pom_paths)

    for idx, ruta_pom in enumerate(pom_paths, 1):
        print(f"\n🔢 Procesando {idx} de {total}: {ruta_pom}")
        ruta_base = ruta_pom.replace("pom.xml", "")

        if modificar:
            carpeta_modificada = copiar_contenido(ruta_base)
            maven_string = update_pom_in_jar(carpeta_modificada, target_pom_version, os.path.basename(ruta_pom))
            if maven_string:
                maven_path = maven_string.replace("pom.xml", "")
                print("🔨 Ejecutando Maven en:", maven_path)
                maven.build_maven_project(maven_path)

        encontrados.append(ruta_pom)

    return encontrados


def extraer_hasta_v(nombre):
    match = re.search(r'^.*?-v(\d{2})\b', nombre)
    return match.group(0) if match else nombre

def generar_reporte(lista_inicial, lista_final, ruta_reporte):
    # Comparar por nombre de archivo base, típicamente pom.xml
    nombres_iniciales = {os.path.relpath(p, constants.SOURCE) for p in lista_inicial}
    nombres_finales = {os.path.relpath(p, constants.MODIFIED) for p in lista_final}

    comunes = nombres_iniciales & nombres_finales
    faltantes = nombres_iniciales - nombres_finales
    nuevos = nombres_finales - nombres_iniciales

    with open(ruta_reporte, 'w', encoding='utf-8') as f:
        f.write("📄 REPORTE DE COMPARACIÓN DE ARCHIVOS\n\n")
        f.write(f"TOTAL INICIALES: {len(nombres_iniciales)}\n")
        f.write(f"TOTAL NUEVA VERSION: {len(nombres_finales)}\n\n")

        f.write("✅ Coinciden (mismo nombre relativo):\n")
        for archivo in sorted(comunes):
            f.write(f"  - {archivo}\n")

        f.write("\n🟥 Faltan en la lista final (no migrados o error):\n")
        for archivo in sorted(faltantes):
            f.write(f"  - {archivo}\n")

        f.write("\n🟩 Nuevos en lista final (no estaban en la original):\n")
        for archivo in sorted(nuevos):
            f.write(f"  - {archivo}\n")

    print(f"✅ Reporte generado en: {ruta_reporte}")


def return_data():
    # Obtener la ruta base (carpeta del script actual)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    PARENT_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))
    def get_row_count(file_name):
        # Buscar archivo en carpeta padre
        file_path = os.path.join(PARENT_DIR, file_name)
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            return wb.active.max_row
        return 0

    tot_num_filas = get_row_count(constants.PRODUCTIVE_XLSX)
    err_num_filas = get_row_count(constants.ERROR_XLSX)
    clo_num_filas = get_row_count(constants.CLONE_ERROR_XLSX)

    data = {
        "total": tot_num_filas,
        "errores": err_num_filas,
        "ok": max(tot_num_filas - err_num_filas, 0),
        "clone": clo_num_filas
    }

    return data


# ------------------- EJECUCIÓN PRINCIPAL -------------------
def execute(version):
    target_pom_version = version
    ruta_inicial = os.getenv("RUTA_INICIAL") or constants.SOURCE
    ruta_final = os.getenv("RUTA_FINAL") or constants.MODIFIED

    if not os.path.exists(ruta_inicial):
        print(f"❌ ERROR: Ruta inicial no existe: {ruta_inicial}")
        sys.exit(1)

    os.makedirs(ruta_final, exist_ok=True)

    lista_inicial = listar_jars_recursivo(target_pom_version,ruta_inicial, modificar=True)
    lista_final = listar_jars_recursivo(target_pom_version,ruta_final, modificar=False)

    generar_reporte(lista_inicial, lista_final, constants.REPORT)

    with open(constants.REPORT, 'r') as f:
        sys.stdout.write(f.read())
    
    return return_data()
