import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook

file_path = "cambios_versiones.xlsx"
ns = {'m': 'http://maven.apache.org/POM/4.0.0'}
new_parent_version = "2.0.6"

# Excel: abrir o crear SIN reescribir encabezado
archivo_existente = os.path.exists(file_path)

if archivo_existente:
    wb = load_workbook(file_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cambios de versión"
    ws.append(["artifactId", "old_version", "new_version", "archivo"])  # encabezado solo si es nuevo

contador = 0

for root_dir, _, files in os.walk("/ruta/a/tu/repositorio"):
    for file in files:
        if file == "pom.xml":
            archivo = os.path.join(root_dir, file)
            print(f"📄 Analizando: {archivo}")

            try:
                tree = ET.parse(archivo)
                root = tree.getroot()

                parent_node = root.find('m:parent', ns)
                artifact_node = root.find('m:artifactId', ns)

                if parent_node is not None and artifact_node is not None:
                    version_node = parent_node.find('m:version', ns)
                    if version_node is not None:
                        old_version = version_node.text
                        version_node.text = new_parent_version

                        # Append escribe en la siguiente fila automáticamente
                        ws.append([
                            artifact_node.text,
                            old_version,
                            new_parent_version,
                            os.path.relpath(archivo)
                        ])
                        contador += 1
                        print(f"✅ Cambiado {artifact_node.text}: {old_version} → {new_parent_version}")
                    else:
                        print("⚠️  <version> no encontrado en <parent>")
                else:
                    print("⚠️  No hay <parent> o <artifactId>")
            except Exception as e:
                print(f"❌ Error procesando {archivo}: {e}")

# Guardar el archivo una sola vez al final
wb.save(file_path)
print(f"\n✅ Excel actualizado con {contador} cambio(s): {file_path}")
