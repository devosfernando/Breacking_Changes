import zipfile
import os
import tempfile
import xml.etree.ElementTree as ET
import shutil
import maven
import re
import sys
import validator_class
from openpyxl import load_workbook, Workbook


target_pom_version = "2.2.0"
file_path = "cambios_versiones.xlsx"

def update_pom_in_jar(jar_path, new_parent_version, filename,  output_path=None):
    with tempfile.TemporaryDirectory() as tmpdir:
        # 1. Extraer el JAR
        with zipfile.ZipFile(jar_path, 'r') as jar:
            jar.extractall(tmpdir)

        # 2. Buscar pom.xml
        pom_path = None
        for root_dir, _, files in os.walk(tmpdir):
            for file in files:
                if file == 'pom.xml':
                    pom_path = os.path.join(root_dir, file)
                    break
        
        if not pom_path:
            print("❌ pom.xml No se encontro el archivo pom en el jar.")
            return

        # 3. Parsear el pom.xml
        tree = ET.parse(pom_path)
        root = tree.getroot()
        ns_uri = root.tag.split('}')[0].strip('{')
        ns = {'m': ns_uri}
        
        # Excel: abrir o crear SIN reescribir encabezado
        archivo_existente = os.path.exists(file_path)

        if archivo_existente:
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cambios de versión"
            ws.append(["artifactId", "old_version", "new_version", "archivo"])  # encabezado solo si es nuevo

        # 4. Cambiar la versión del parent
        parent_node = root.find('m:parent', ns)
        artifact =  root.find('m:artifactId', ns)
        if parent_node is not None:
            parent_version_node = parent_node.find('m:version', ns)
            if parent_version_node is not None:
                print(f"🔧 Cambiando versión del parent: {parent_version_node.text} → {new_parent_version} job → {artifact.text}")
                 # Agregar fila al Excel
                ws.append([artifact.text, parent_version_node.text, new_parent_version])
                parent_version_node.text = new_parent_version
            else:
                print("⚠️ <parent><version> not found.")
        else:
            print("⚠️ <parent> section not found.")

        # 6. Guardar pom modificado
        ET.register_namespace('', ns_uri)
        tree.write(pom_path, encoding="utf-8", xml_declaration=True)

        
        carpeta_salida = "./modified"
        
        os.makedirs(carpeta_salida, exist_ok=True)
        # Guardar Excel
        wb.save("cambios_versiones.xlsx")
        #output_path = os.path.join(carpeta_salida, filename)

        guardar_archivos_en_carpeta(tmpdir,filename)

        #print(f"✅ JAR actualizado exitosamente: {filename}")
        return pom_path


def guardar_archivos_en_carpeta(tmpdir, jar_name):
    nombre_base = os.path.splitext(jar_name)[0]
    output_dir = os.getenv("RUTA_FINAL", "modified")
    destino_base = os.path.join(output_dir, nombre_base)
    os.makedirs(destino_base, exist_ok=True)

    for root_dir, _, files in os.walk(tmpdir):
        for file in files:
            full_path = os.path.join(root_dir, file)
            relative_path = os.path.relpath(full_path, tmpdir).replace("\\", "/")
            destino_final = os.path.join(destino_base, relative_path)

            os.makedirs(os.path.dirname(destino_final), exist_ok=True)
            shutil.copy2(full_path, destino_final)

def listar_jars_recursivo(directorio, isInit):

    if not os.path.isdir(directorio):
        print(f"Directorio no válido: {directorio}")
        return

    jars_encontrados = []
    # Contar primero
    jar_count = 0
    jar_sig = 0
    for root, _, files in os.walk(directorio):
        for file in files:
            if file.endswith(".jar"):
                jar_count += 1
    for root, _, files in os.walk(directorio): #se utiliza el os.walk para que busque en las subcarpetas
        for idx, file in enumerate(files):
            jar_sig += 1
        #for file in files:
            if isInit:
                print(f"\U0001F552 Ejecutando : {jar_sig} → {jar_count}")
            if file.endswith(".jar"):
                ruta_completa = os.path.join(root, file).replace("\\", "/") #se reemplazan los back slash por slash
                if isInit:               
                    ruta_maven = file.replace('.jar','')
                    pom_path = update_pom_in_jar(ruta_completa,target_pom_version,file)
                    before, sep, after = pom_path.partition("META-INF")

                    maven_path =  './modified/{name}/META-INF{after}'.format(name=ruta_maven,after=after).replace("\\", "/")
                    maven_path = maven_path.replace("pom.xml","")

                    maven.build_maven_project(maven_path)
                jars_encontrados.append(ruta_completa)
    return jars_encontrados

def extraer_hasta_v(nombre):  # se utiliza esta expresion ya que los jar de artifactory vienen con hash
    # Extrae hasta la primera ocurrencia de "-v"
    match = re.search(r'^.*?-v(\d{2})\b', nombre)
    return match.group(0) if match else nombre

def generar_reporte(lista_inicial, lista_final, ruta_reporte):
    # Extraer nombres hasta '-v'
    iniciales = {extraer_hasta_v(os.path.basename(p)) for p in lista_inicial}
    finales = {extraer_hasta_v(os.path.basename(p)) for p in lista_final} 

    comunes = iniciales & finales
    faltantes = iniciales - finales
    nuevos = finales - iniciales
    


    with open(ruta_reporte, 'w', encoding='utf-8') as f:
        total_ini = len(iniciales)        
        total_com = len(finales)

        f.write("====================================📄 REPORTE DE COMPARACIÓN DE ARCHIVOS ====================================\n\n")
        f.write("=================TOTAL INICIALES:      "+str(total_ini)+"\n\n")
        f.write("=================TOTAL NUEVA VERSION:  "+str(total_com)+"\n\n")

        f.write("✅ Coinciden:\n")
        for archivo in sorted(comunes):
            f.write(f"  - {archivo}\n")
        f.write("\n")

        f.write("🟥 Faltan en lista final (están en inicial):\n")
        for archivo in sorted(faltantes):
            f.write(f"  - {archivo}\n")
        f.write("\n")

        f.write("🟩 Nuevos en lista final (no estaban antes):\n")
        for archivo in sorted(nuevos):
            f.write(f"  - {archivo}\n")

    print(f"✅ Reporte generado en: {ruta_reporte}")

#lista_inicial = listar_jars_recursivo("E:/BBVA/LRBA/TEST-CABECERAS/Script_Artefactory/Peru_Descargas_Atifactory", True)
#lista_final = listar_jars_recursivo("./modified", False)
ruta_inicial = os.getenv("RUTA_INICIAL") or "./source"
ruta_final = os.getenv("RUTA_FINAL") or "/modified"
if not ruta_inicial or not ruta_final:
    print("❌ ERROR: Variables de entorno RUTA_INICIAL o RUTA_FINAL no están definidas.")
    sys.exit(1)


if not os.path.exists(ruta_final):
    print(f"ℹ️  Creando directorio: {ruta_final}")
    os.makedirs(ruta_final, exist_ok=True)

lista_inicial = listar_jars_recursivo(ruta_inicial, True)
lista_final = listar_jars_recursivo(ruta_final, False)

generar_reporte(lista_inicial, lista_final, "reporte_comparacion.txt")
validator_class.check_maven_project("./modified")
with open("./reporte_comparacion.txt", 'r' ) as file:
    sys.stdout.write(file.read())